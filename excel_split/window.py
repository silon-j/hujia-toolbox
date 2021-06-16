import sys, os, time, logging, datetime
from traceback import print_tb
from PyQt5 import QtCore, QtWidgets
import PyQt5
from PyQt5.QtGui import QPixmap, QPainter, QColor, QFont, QIcon
from PyQt5.QtWidgets import (QApplication, QComboBox, QMainWindow, QFileDialog, QMessageBox, 
QAction, QProgressBar, QPushButton, QVBoxLayout, QGridLayout, QHBoxLayout, QWidget, QListView, QLineEdit, QLabel)
from openpyxl.cell import cell

from win32com import client as wc
from openpyxl import Workbook, load_workbook
from utility import *



class MainWindow(QMainWindow):


    def __init__(self):
        super().__init__()
        self.init_ui()
        self.connect_action()

    def init_ui(self):
        self.setWindowTitle('excel 拆分工具')
        self.setStyleSheet('#fileTypeTrans{background-color:white}')
        self.setWindowIcon(QIcon('../static/img/logo-circle.png'))
        self.setMinimumSize(480, 300)
        self.mainLayout = QGridLayout()

        # main content
        self.centralWidget = QWidget(self)
        self.centralWidget.setObjectName('centralWidget')
        self.centralWidget.setLayout(self.mainLayout)
        self.setCentralWidget(self.centralWidget)



class SplitWidget(QWidget):
    inputFileName = ''
    inputFilePath = ''
    fileHeaders = []
    headerSelected = ''
    headerSelectedIndex = None
    headerColSet = []
    autoColWidth = True
    startTime = None
    endTime = None

    def __init__(self):
        super().__init__()
        self.init_ui()
        self.connect_action()

    def init_ui(self):
        # self.centralWidget = QtWidgets.QWidget(self)
        # self.centralWidget.setObjectName("centralWidget")
        self.setWindowTitle('excel 表格拆分')
        self.setWindowIcon(QIcon('../static/img/logo-circle.png'))
        self.setMinimumSize(400,300)
        self.mainLayout = QGridLayout()

        # input file path button
        self.inputFileBtn = QPushButton()
        self.inputFileBtn.setText('选择需要拆分的 xlsx 文件')


        # input file box
        self.filePicLabel = QLabel()
        self.filePic = QPixmap("../static/img/excel.png")
        self.filePic_scaled = self.filePic.scaled(100, 100)
        self.filePicLabel.setPixmap(self.filePic_scaled)
        self.filePicLabel.setAlignment(QtCore.Qt.AlignCenter)

        # file name label
        self.fileNameLabel = QLabel()
        self.fileNameLabel.setText('未选择文件')
        self.fileNameLabel.setWordWrap(True)
        self.fileNameLabel.setAlignment(QtCore.Qt.AlignCenter)

        # start split btn
        self.startBtn = QPushButton()
        self.startBtn.setText('开始')

        # reset btn
        self.resetBtn = QPushButton()
        self.resetBtn.setText('重置')

        # header filter select
        self.headerSelect = QComboBox()
        self.headerSelect.setView(QListView())
        self.headerSelect.setObjectName('headerSelect')
        self.headerSelect.setGeometry(QtCore.QRect(60, 60, 260, 28))
        self.headerSelect.setStyleSheet('''
            #headerSelect
            {
                background: #eeeeee;
                padding-left:5px ;
                border-radius:3px;
                border: 1px solid gray;
            }
            #headerSelect QAbstractItemView::item
            {
                height:40px;
            }
        ''')

        # process bar
        self.processBar = QProgressBar()
        self.processBar.setProperty("value", 0)

        self.mainLayout.addWidget(self.inputFileBtn,3,0,1,6)
        self.mainLayout.addWidget(self.filePicLabel,1,0,1,6)
        self.mainLayout.addWidget(self.fileNameLabel,2,0,1,6)
        self.mainLayout.addWidget(self.headerSelect,4,0,2,4)
        self.mainLayout.addWidget(self.startBtn,4,4,2,2)
        self.mainLayout.addWidget(self.processBar,5,0,2,6)

        self.setLayout(self.mainLayout)

    def connect_action(self):
        self.inputFileBtn.clicked.connect(self.set_inputfile)
        self.startBtn.clicked.connect(self.start_split)
        self.headerSelect.currentIndexChanged[str].connect(self.select_header)

    def set_inputfile(self):
        self.inputFilePath, filter = QFileDialog.getOpenFileName(
            self, '选择待处理文件', r'C:\Users\Administrator\Desktop', 'excel 文件(*.xlsx)')
        self.inputFileName = os.path.basename(self.inputFilePath)
        self.fileNameLabel.setText(self.inputFilePath)

        wb = load_workbook(self.inputFilePath)
        ws = wb.worksheets[0]
        self.ws = ws
        self.rowCnt = ws.max_row
        self.fileHeaders = self.get_row_values(ws=ws, row_num=1)
        self.headerSelect.addItems(self.fileHeaders)

    # TODO: 可拆分至excel class中
    def get_row_values(self, ws, row_num):
        colCnt = ws.max_column
        rowData = []
        for i in range(1, colCnt + 1):
            v = ws.cell(row=row_num, column=i).value
            rowData.append(v)
        return rowData

    def select_header(self, header):
        self.headerSelected = header
        self.headerSelectedIndex = self.fileHeaders.index(header)
        self.headerColSet = self.get_col_set(ws=self.ws, colNum=self.headerSelectedIndex + 1)

    def init_ui_status(self):
        pass

    def set_ui_enabled(self, flag):
        self.inputFileBtn.setEnabled(flag)
        self.headerSelect.setEnabled(flag)
        self.startBtn.setEnabled(flag)


    def start_split(self):
        if self.inputFilePath:
            check_dir('./Excel表拆分')
            self.set_ui_enabled(False)
            self.startTime = datetime.datetime.now()
            self.split_excel()

        else:
            QMessageBox.warning(self, '提示','请选择待处理文件')

    def get_col_set(self, ws, colNum):
        dataSet = set()
        rowCnt = self.rowCnt
        for i in range(2, rowCnt + 1):
            v = ws.cell(row=i, column=colNum).value
            dataSet.add(v)
        QMessageBox.about(self, '提示', '该关键词下共有{count}种值'.format(count=len(dataSet)))
        # print(dataSet)
        return dataSet


    def split_excel(self):
        errorCnt = 0
        strStartTime = str(self.startTime)[0:19].replace(':', '-')
        # create child base excel file
        for i, v in enumerate(self.headerColSet):
            locals()['wb_' + str(i)] = Workbook()
            locals()['ws_' + str(i)] = locals()['wb_' + str(i)].active

            s = locals().get('ws_' + str(i))
            f = locals().get('wb_' + str(i))

            # TODO: 通用化配置，例如统一添加自定义添加表头
            check_dir('./Excel表拆分/{time_now}'.format(time_now=strStartTime))
            f.save('./Excel表拆分/{time_now}/{file_name}'
                   .format(time_now=strStartTime, file_name=str(v) + '.xlsx'))

        # map parent excel rows, check sp key, copy rows to child excel
        for rowNum in range(1, self.rowCnt + 1):
            rowData = self.get_row_values(ws=self.ws, row_num=rowNum)
            for i, v in enumerate(self.headerColSet):
                try:
                    if rowData[self.headerSelectedIndex - 1] == v:
                        s = locals().get('ws_' + str(i))
                        f = locals().get('wb_' + str(i))
                        s.append(rowData)
                        f.save('./Excel表拆分/{time_now}/{file_name}'
                            .format(time_now=strStartTime, file_name=str(v) + '.xlsx'))
                except Exception as e:
                    print(e)
                    errorCnt += 1
                self.processBarValue = rowNum * 1 / self.rowCnt * 100
                self.processBar.setValue(self.processBarValue)
        self.endTime = datetime.datetime.now()
        QMessageBox.about(self, '处理报告', '处理完成\n用时{}\n共处理{}行,失败{}行'.format(
            self.endTime-self.startTime, self.rowCnt, errorCnt))
        self.set_ui_enabled(True)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main = SplitWidget()
    main.show()
    sys.exit(app.exec_())