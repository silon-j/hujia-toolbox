#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@File   : app.py
@CreateTime   : 2019/9/19 17:57
@Version: 0.3a
@Author : silon-j
@Desc   : None
"""

# import lib
import os
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.styles.colors import RED, BLUE, YELLOW, GREEN, BLACK, WHITE
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import *

import datetime

# const
color_list = ['RED', 'BLUE', 'YELLOW', 'GREEN', 'BLACK', 'WHITE']

1

def fn_select():
    show_tip_block('互加计划Excel处理工具')
    fn_number = input(
        '请输入要使用的功能数字\n'
        '1. 拆分Excel表\n'
        '2. 批量添加表头（即将推出）\n'
        '3. 批量自动调整列宽（即将推出）\n'
        '4. 开发中\n'
        '0. 退出\n'
        + '/' * 40 + '\n'
    )
    if fn_number == '1':
        excel_split()
    elif fn_number == '2':
        pass
    elif fn_number == '0':
        pass
    else:
        print('\n>>>输入类型错误，请重新尝试~\n')
        fn_select()


def show_tip_block(content, width=40):
    c_width = int(get_string_cn_width(content))
    print('\n' + ('+' * width))
    print('|' + content.center(width - c_width - 2) + '|')
    print('+' * width)


def get_string_cn_width(s):
    length = len(s)
    # print('length:', length)
    utf8_length = len(s.encode('utf-8'))
    # print('utf-8 length:', utf8_length)
    return (utf8_length - length) / 2


def get_string_full_width(s):
    return len(s.encode('utf-8')) / 3 + len(s)


def load_excel_sheet(f_name, sheet_idx=0):
    global wb, ws
    wb = load_workbook('./{f_name}'.format(f_name=f_name))
    ws = wb.worksheets[sheet_idx]
    return ws


def get_sp_col_idx(sp_key):
    r_vls = get_row_values(1, ws)
    if sp_key in r_vls:
        idx = r_vls.index(sp_key)
        print('\n成功选取！当前筛选列为第{count}列'.format(count=idx + 1))
        return idx + 1
    else:
        print(
            '未找到筛选列！可筛选关键词如下\n'
            '{v}'.format(v=r_vls)
        )


def get_row_values(row_count, worksheet):
    columns = worksheet.max_column
    row_data = []
    for i in range(1, columns + 1):
        v = worksheet.cell(row=row_count, column=i).value
        row_data.append(v)
    return row_data


def get_sp_col_val_set(col_count):
    print('筛选列数当前为{col_count}\n'.format(col_count=col_count))
    row = ws.max_row
    col_data_set = set()
    for i in range(2, row + 1):
        v = ws.cell(row=i, column=col_count).value
        col_data_set.add(v)

    print('该关键词下共有{count}种值'.format(count=len(col_data_set)))
    print(col_data_set)

    return col_data_set


def line_divider(s='/', count=40):
    print('\n')
    print(s * count)


def add_header():
    line_divider()
    add = input('是否为子表添加表头？（y/n）\n').lower()
    if add == 'y':
        return get_header_info_list()
    elif add == 'n':
        return []


def get_header_info_list():
    header = []
    for i in range(1, 1001):
        # get content settings
        hd_r = dict()
        hd_r['row'] = i
        hd_r['content'] = input('请输入第{i}行内容\n'.format(i=i))
        hd_r['center'] = (True if input('是否居中(y/n)\n').lower() == 'y' else False)
        hd_r['font_size'] = sel_font_size(element_name='单元格', default_size=16)
        hd_r['font_color'] = sel_color(element_name='字体')
        hd_r['bg_color'] = sel_color(element_name='单元格背景')
        header.append(hd_r)

        # if go on add header ?
        line_divider()
        go = input('是否继续添加表头？(y/n)\n')
        if go == 'y':
            pass
        elif go == 'n':
            break

    return header


def render_cell(sheet, row, col, content, font_size, font_color, center, bg_color):
    # render cell
    cell = sheet.cell(row=row, column=col, value=content)
    cell.font = Font(name='等线', size=font_size, color=font_color)
    cell.alignment = Alignment(horizontal=('center' if center else 'left'), vertical='center')
    cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)


def sel_font_size(element_name, default_size=16):
    print('\n请输入{element}字号(数字) \n回车 默认设置为 {default_size}'
          .format(element=element_name, default_size=default_size))
    size = input()
    if size == '':
        return default_size
    else:
        try:
            font_size = int(size)
            return font_size
        except Exception as e:
            print('请正确输入字号!错误如下：\n' + str(e))
            return sel_font_size(element_name, default_size=16)


def sel_color(element_name, default_color='BLACK', default_rgb='D9E1F2'):
    print('\n请输入{element}颜色，以下可选：\n'
          'RED, BLUE, YELLOW, GREEN, BLACK, WHITE \n'
          '输入RGB 进入RGB颜色选择模式\n'
          '回车 默认设置为 {default_color}'
          .format(element=element_name, default_color=default_color))
    color = input().upper()
    if color in color_list:
        return eval(color)
    elif color == '':
        return BLACK
    elif color == 'RGB':
        print('请输入{element}rgb颜色 HEX 值. 例： D9E1F2 \n'
              '请严格按照示例格式输入！\n'
              '输入back返回上一步颜色选择\n'
              '回车 默认设置为 D9E1F2 浅蓝色'
              .format(element=element_name, default_rgb=default_rgb))
        rgb = input()
        if rgb == '':
            return default_rgb
        elif rgb == 'back':
            return sel_color(element_name,
                             default_color='BLACK',
                             default_rgb='D9E1F2')
        else:
            return rgb


def merge_row(sheet, row_count):
    # get max col letter
    max_col_letter = get_column_letter(ws.max_column)

    # insert row at row_count
    sheet.insert_rows(row_count)

    # merge row cells
    sheet.merge_cells('{c_l}:{c_r}'
                      .format(c_l='A' + str(row_count),
                              c_r=max_col_letter + str(row_count)
                              )
                      )


def excel_split():
    # open  excel work sheet
    load_excel_sheet(get_file_name())

    # key word for table filter
    sp_key = input('请输入拆表关键词\n')

    sp_col_count = get_sp_col_idx(sp_key)
    sp_col_data_set = get_sp_col_val_set(sp_col_count)

    start_time = datetime.datetime.now()
    str_start_time = str(start_time)[0:19].replace(':', '-')
    check_dir('Excel表拆分/{time_now}'
              .format(time_now=str_start_time))

    # add header
    header = add_header()

    #
    print('拆表开始...')

    # create child excel file
    for i, v in enumerate(sp_col_data_set):
        locals()['wb_' + str(i)] = Workbook()
        locals()['ws_' + str(i)] = locals()['wb_' + str(i)].active

        s = locals().get('ws_' + str(i))
        f = locals().get('wb_' + str(i))

        for hd_val in header:
            merge_row(sheet=s, row_count=hd_val['row'])
            render_cell(sheet=s, row=hd_val['row'], col=1,
                        content=hd_val['content'],
                        font_size=hd_val['font_size'],
                        font_color=hd_val['font_color'],
                        center=hd_val['center'],
                        bg_color=hd_val['bg_color'])
        s.append(get_row_values(1, ws))
        f.save('./Excel表拆分/{time_now}/{file_name}'
               .format(time_now=str_start_time,
                       file_name=str(v) + '.xlsx'))

    # map parent excel rows, check sp key, copy rows to child excel
    for row_count in range(1, ws.max_row + 1):
        a = get_row_values(row_count, ws)
        print('row: ', row_count, a)
        row = get_row_values(row_count, ws)
        for i, v in enumerate(sp_col_data_set):
            if row[sp_col_count - 1] == v:
                # print('value match')
                s = locals().get('ws_' + str(i))
                f = locals().get('wb_' + str(i))
                s.append(a)
                f.save('./Excel表拆分/{time_now}/{file_name}'
                       .format(time_now=str_start_time,
                               file_name=str(v) + '.xlsx'))

    for i, v in enumerate(sp_col_data_set):
        s = locals().get('ws_' + str(i))
        f = locals().get('wb_' + str(i))

        auto_col_width(sheet=s, begin_row_index=len(header))
        f.save('./Excel表拆分/{time_now}/{file_name}'
               .format(time_now=str_start_time,
                       file_name=str(v) + '.xlsx'))

    end_time = datetime.datetime.now()
    show_process_time_report(start_time, end_time)
    fn_select()


def auto_col_width(sheet, begin_row_index=3):
    print('\n正在自动调整表格列宽...')
    col_width = []
    # map col
    for i, col in enumerate(sheet.columns):
        # map row value count in this col
        for row_count in range(begin_row_index, len(col)):
            cell_width = get_string_full_width(str(col[row_count].value))
            if row_count == begin_row_index:
                col_width.append(0)
            else:
                # get max width in each col
                if col_width[i] < cell_width:
                    col_width[i] = cell_width

    # set col width
    for i in range(len(col_width)):
        col_letter = get_column_letter(i + 1)
        # set col width
        sheet.column_dimensions[col_letter].width = col_width[i] + 2

    print('调整完毕！')


def check_dir(dir_name):
    if not os.path.exists("./{name}".format(name=dir_name)):
        os.makedirs('{name}'.format(name=dir_name))


def show_process_time_report(start, end):
    show_tip_block('文件处理报告')
    print('开始时间：{start_time}'.format(start_time=start))
    print('完成时间：{end_time}'.format(end_time=end))
    print('处理用时：{process_time}\n'.format(process_time=(end - start)))
    return


def get_file_name():
    file_name = input("请输入excel文件名（含拓展名）：\n")
    # file_name = str(a) + ".xlsx"
    if os.path.exists("./" + file_name):
        return file_name
    else:
        print("该excel文件不存在，请重新输入文件名称")
        return get_file_name()


def check_input_value(input_content, check_type, fn):
    try:
        return check_type(input_content)
    except Exception as e:
        print('\n>>>输入类型错误，请重新尝试~\n'
              '错误如下：{e}\n'.format(e=e))
        eval(fn())
        return check_input_value(input_content, check_type, fn)


if __name__ == '__main__':

    fn_select()
